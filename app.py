import streamlit as st
import openpyxl
from io import BytesIO
from html import escape

st.set_page_config(layout="wide")

st.title("Hello Sales Team 👋 Intamarque Offer Sheet to Brevo HTML Converter APP")
st.write("Upload your Excel offer sheet and this will generate your html code with css inline.")

uploaded_file = st.file_uploader("📤 Upload Excel File", type=["xlsx"])

# Convert Excel fill color to hex with theme fallback
def get_bg_color(cell):
    try:
        fill = cell.fill
        if fill.patternType == 'solid':
            if fill.fgColor.type == 'rgb' and fill.fgColor.rgb:
                rgb = fill.fgColor.rgb
                if len(rgb) == 8:
                    return f"#{rgb[2:]}"
            elif fill.fgColor.type == 'theme':
                theme_colors = {
                    0: "#ffffff",  # Light1
                    1: "#000000",  # Dark1
                    2: "#eeece1",  # Light2
                    3: "#1f497d",  # Dark2
                }
                return theme_colors.get(fill.fgColor.theme, "#ffffff")
    except:
        pass
    return "#ffffff"

# Bold detection
def is_bold(cell):
    try:
        return cell.font.bold
    except:
        return False

# Format values neatly
def format_value(val, number_format):
    if val is None:
        return ""
    try:
        if "£" in number_format or "\u00a3" in number_format:
            return f"£{float(val):,.2f}"
        elif "$" in number_format:
            return f"${float(val):,.2f}"
        elif "€" in number_format or "\u20ac" in number_format:
            return f"€{float(val):,.2f}"
        elif isinstance(val, float):
            return f"{val:,.2f}"
        return str(val)
    except:
        return escape(str(val))

# Main table builder with scroll + fixed width
def generate_html(sheet):
    max_col_with_data = max(
        (cell.column for row in sheet.iter_rows() for cell in row if cell.value not in [None, ""]),
        default=0
    )

    html = '''
    <div style="overflow-x: auto; width: 100%;">
        <table style="border-collapse: collapse; font-family: Arial, sans-serif; font-size: 12px; width: 1500px; table-layout: fixed;">
    '''
    for i, row in enumerate(sheet.iter_rows(min_row=6), start=6):
        if all(cell.value in [None, ""] for cell in row[:max_col_with_data]):
            continue

        html += "<tr>"
        for j, cell in enumerate(row[:max_col_with_data]):
            value = format_value(cell.value, cell.number_format)

            # Custom colours for specific columns
            if j == 9 or j == 10:  # Columns J and K
                bg = "#ffe5cc"
            elif j == 11 or j == 12:  # Columns L and M
                bg = "#e6f4e6"
            else:
                bg = get_bg_color(cell)

            bold = "font-weight: bold;" if is_bold(cell) else ""
            align = "text-align: center;" if isinstance(cell.value, (int, float)) else "text-align: left;"
            html += f'<td style="border: 1px solid #ccc; padding: 6px; background-color: {bg}; white-space: nowrap; {bold} {align}">{value}</td>'
        html += "</tr>"
    html += "</table></div>"
    return html

if uploaded_file:
    wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()), data_only=True)
    sheet = wb.active
    html_code = generate_html(sheet)

    st.subheader("✅ Brevo-Ready HTML")
    st.text_area("👇 Copy this into your Brevo HTML block:", html_code, height=500)
    st.success("All done — styling and spacing now match Excel perfectly! 🚀")
